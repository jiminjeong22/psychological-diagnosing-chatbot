from openpyxl import load_workbook
from konlpy.tag import Komoran


class Sentiment_Analysis:
    def __init__(self):
        with load_workbook(filename='../data/polarity.xlsx')["Worksheet"] as self.sentiment_file:
            # dic posses all sentiment words
            self.sentiments = {}
            for row_line in range(2, self.sentiment_file.max_row):
                text = str(self.sentiment_file.cell(row=row_line, column=1).value)
                text = text.split(";")
                if text[0] not in self.sentiments:
                    self.sentiments[text[0]] = [row_line]
                else:
                    self.sentiments[text[0]].append(";".join(text[1:]))

    def calculate_result(self, text):
        komoran = Komoran()
        pos_l = list(komoran.pos(text))

        # change format to compare with polarity.xlsx
        for i in range(len(pos_l)):
            pos_l[i] = "/".join(list(pos_l[i]))

        POS = 0
        NEG = 0

        l = []

        i = 0

        while i < len(pos_l):
            if pos_l[i] in self.sentiments:
                text = pos_l[i]
                if i + 1 < len(pos_l):
                    if pos_l[i + 1] in self.sentiments[pos_l[i]]:
                        text = pos_l[i] + ";" + pos_l[i + 1]
                        i += 1
                        if i + 1 < len(pos_l):
                            if pos_l[i] + ";" + pos_l[i + 1] in self.sentiments[pos_l[i - 1]]:
                                text = pos_l[i - 1] + ";" + pos_l[i] + ";" + pos_l[i + 1]
                                index = self.sentiments[pos_l[i - 1]][0] + self.sentiments[pos_l[i - 1]].index(
                                    pos_l[i] + ";" + pos_l[i + 1])
                                i += 1
                            else:
                                index = self.sentiments[pos_l[i - 1]][0] + self.sentiments[pos_l[i - 1]].index(pos_l[i])
                        else:
                            index = self.sentiments[pos_l[i - 1]][0] + self.sentiments[pos_l[i - 1]].index(pos_l[i])
                    else:
                        index = self.sentiments[pos_l[i]][0]
                else:
                    index = self.sentiments[pos_l[i]][0]

                neg = int(float(self.sentiment_file.cell(row=index, column=2).value) * 100) / 100
                pos = int(float(self.sentiment_file.cell(row=index, column=3).value) * 100) / 100
                POS += pos
                NEG += neg
                l.append((text, pos, neg))
            i += 1

        POS = int(POS * 100) / 100
        NEG = int(NEG * 100) / 100
        mood = (POS + NEG)

        print("CURRENT MOOD IS POS:{},NEG:{}, TOTAL:{}".format(POS, NEG, mood))

        return pos_l, POS, NEG, mood
