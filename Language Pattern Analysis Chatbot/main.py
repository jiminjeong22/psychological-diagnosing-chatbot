import json
import random
from openpyxl import *
from flask import Flask, request, make_response, jsonify
from datetime import datetime
import gensim
from eunjeon import Mecab
import re
 
app = Flask(__name__)
log = app.logger
data = []

m=Mecab()
model = gensim.models.Doc2Vec.load("Handong1.model")

from openpyxl import *
wb = load_workbook("./Data/"+"corpus.xlsx")
ws= wb.get_sheet_by_name("All")
from eunjeon import Mecab
m=Mecab()
import gensim

from soynlp.word import WordExtractor
from soynlp.tokenizer import LTokenizer

def corpus_making():
    corpus = {}
    row=2
    while (ws.cell(row=row, column =3)).value >100:
        if (ws.cell(row=row, column =2)).value in pos_l:
            corpus[(ws.cell(row=row, column =1)).value] = ((ws.cell(row=row, column =2)).value,(ws.cell(row=row, column =3)).value) #(Pos,Frequency)
        row+=1
    print(len(corpus))
    return(corpus)
    
pos_l=['NNG', 'NNP', 'NP', 'MAG', 'MAJ', 'VCN', 'VA']
corpus= corpus_making()


@app.route('/', methods=['POST','GET'])
def webhook():
    
    req = request.get_json(silent=True, force=True)
    #print(req)
    try:
        action = req.get('queryResult').get('action')
        intent = req['queryResult']['intent']['displayName']
           
    except AttributeError:
        return 'json error'
    
    if action == 'input.welcome':
        res = "진단 시작"
 
    elif action == 'input.unknown':
        text = str(req.get('queryResult').get('queryText'))
        print(text)
        word_similarity = word_frequency_diagnosis(corpus, text)
        context_similarity= Context_Similarity_Diagnosis(text)
        res= word_similarity+"\n"+context_similarity      
                        
    else:
        log.error('Unexpected action.')
   
    print("*"*20)
    print("\n")
    return make_response(jsonify({'fulfillmentText': res}))
    

def Context_Similarity_Diagnosis(given_text):
    similarity = 0
    input_sentence = given_text
    doc_list = m.morphs(input_sentence)
    inferred_vector = model.infer_vector(doc_list)
    return_vecs = model.docvecs.most_similar(positive=[inferred_vector],topn=5)
    print(return_vecs)
    for vec in return_vecs:
        similarity+=vec[1]
    similarity = int(similarity/5*1000)/10
    if return_vecs[0][1]<0.76: 
        similarity = similarity-int(30)
    print(("Context similarity: {}% to Handong model".format(similarity)))
    return("Context similarity: {}% to Handong model".format(similarity))



def word_frequency_diagnosis(corpus, given_text):
    similarity = 0
    mean_pos = []
    text_l = m.pos(given_text) #[('요즘', 'MAG'), ('시험', 'NNG'), ('기간', 'NNG'), ('이', 'VCP'), ('라', 'EC'), ('너무', 'MAG'), ('힘들', 'VA'), ('어', 'EC')]
    print(text_l)
    result_l=[]
    for text in text_l:
        if text[1] in pos_l: 
            mean_pos.append(text)
            if text[0] in corpus.keys():
                result_l.append((text[0], text[1], corpus[text[0]][1]))
                similarity+= (int(corpus[text[0]][1])/100)
    print("meaning morpheme count: {}".format(len(mean_pos)))
    print("meaning morphemes: ")
    print(result_l)
    print("meaning morpheme portion: {}%".format(len(result_l)/len(mean_pos)*100))
    print("Meaning Word similarity: {}% to Handong model".format(len(result_l)/len(mean_pos)*100))
    return("Meaning Word similarity: {}% to Handong model".format(len(result_l)/len(mean_pos)*100))
       
if __name__ == '__main__':
    app.run(debug=True)
