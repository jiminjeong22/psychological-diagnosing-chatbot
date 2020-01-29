from openpyxl import load_workbook


class Corpus:
    def __init__(self):
        self.wb = load_workbook("./data/" + "corpus.xlsx")
        self.ws = self.wb.get_sheet_by_name("All")
        self.pos_l = ['NNG', 'NNP', 'NP', 'MAG', 'MAJ', 'VCN', 'VA']

    def corpus_making(self):
        corpus = {}
        row = 2
        while (self.ws.cell(row=row, column=3)).value > 100:
            if (self.ws.cell(row=row, column=2)).value in self.pos_l:
                corpus[(self.ws.cell(row=row, column=1)).value] = (
                    (self.ws.cell(row=row, column=2)).value, (self.ws.cell(row=row, column=3)).value)  # (Pos,Frequency)
            row += 1
        print(len(corpus))
        return corpus
