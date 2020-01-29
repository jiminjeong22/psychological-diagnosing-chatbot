from wordcloud import WordCloud
from openpyxl import *

wb = load_workbook("./Data/"+"corpus.xlsx")
ws1= wb.get_sheet_by_name("KakaoChat")
ws2= wb.get_sheet_by_name("Everytime")
kakao={}
everytime={}
for row in range(2,300):
    if (ws1.cell(row=row, column =2)).value == 'NNG':
        kakao[(ws1.cell(row=row, column =1)).value] = (ws1.cell(row=row, column =3)).value
    if (ws2.cell(row=row, column =2)).value == 'NNG':
        everytime[(ws2.cell(row=row, column =1)).value] = (ws2.cell(row=row, column =3)).value
print(kakao)
font_path = "C:/Windows/Fonts/NanumGothic.ttf"
wordcloud1 = WordCloud(
    background_color="white",
    font_path = font_path,
    width = 800,
    height = 800
    )
wordcloud2 = WordCloud(
    background_color="white",
    font_path = font_path,
    width = 800,
    height = 800)
wordcloud1 = wordcloud1.generate_from_frequencies(kakao)
wordcloud2 = wordcloud2.generate_from_frequencies(everytime)

array1 = wordcloud1.to_array()
array2 = wordcloud2.to_array()

import matplotlib.pyplot as plt

fig = plt.figure(figsize=(10, 10))
plt.imshow(array1, interpolation="bilinear")
plt.axis("off")
plt.show()
fig.savefig('wordcloud_kakao.png')

plt.imshow(array2, interpolation="bilinear")
plt.axis("off")
plt.show()
fig.savefig('wordcloud_everytime.png')